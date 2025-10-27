import openpyxl
import sys
import io

# Fix Windows encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('LSZ__fjfjfe.xlsm', data_only=True)
print("Sheet names:", wb.sheetnames)
print()

ws = wb['Celkové výsledky']

print("=== TABULKA 3 (radek 23+3=26) ===")
for i, row in enumerate(ws.iter_rows(min_row=26, max_row=30, min_col=2, max_col=9, values_only=True), start=26):
    print(f"Radek {i}: {list(row)}")

print("\n=== TABULKA 4 (radek 43+3=46) ===")
for i, row in enumerate(ws.iter_rows(min_row=46, max_row=50, min_col=2, max_col=9, values_only=True), start=46):
    print(f"Radek {i}: {list(row)}")

print("\n=== TABULKA 5 (radek 52+3=55) ===")
for i, row in enumerate(ws.iter_rows(min_row=55, max_row=60, min_col=2, max_col=10, values_only=True), start=55):
    print(f"Radek {i}: {list(row)}")
