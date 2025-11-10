"""
PP Excel Results Reader - načte výsledky z PP Excel a uloží do JSON
Usage: python read_pp_results.py <pp_excel_file>

PP (Pracovní polohy) Excel obsahuje data o polohách těla během práce.
"""
import sys
import json
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional


# =============================================================================
# CONSTANTS - PP Excel Structure
# =============================================================================

# List name
PRUMER_SHEET = "Průměr"

# Column indices (0-based)
COL_NAZEV_POLOHY = 0  # A: Název polohy
COL_TYP_PRACE = 1      # B: Typ svalové práce (Statická/Dynamická)
COL_VYSKYT_A = 2       # C: Výskyt minut pracovník A
COL_VYSKYT_B = 3       # D: Výskyt minut pracovník B
COL_PRUMER = 4         # E: Průměr minut
COL_TYP_POLOHY = 5     # F: Typ pracovní polohy (N/PP)

# Row ranges for each body part section
SECTIONS = {
    "trup": (4, 14),         # Trup: řádky 4-14 (11 položek)
    "hlava_krk": (16, 25),   # Hlava a krk: 16-25 (10 položek)
    "phk": (27, 36),         # PHK: 27-36 (10 položek)
    "lhk": (38, 47),         # LHK: 38-47 (10 položek)
    "dk": (49, 55),          # Dolní končetiny: 49-55 (7 položek)
    "ostatni": (57, 60)      # Ostatní části těla: 57-60 (4 položek)
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def convert_to_json_safe(value: Any) -> Any:
    """Převede hodnotu na JSON-safe formát"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return str(value)


def read_section_data(ws, start_row: int, end_row: int) -> Dict[str, Dict[str, Any]]:
    """
    Načte data z jedné sekce (např. Trup, Hlava a krk, atd.)

    Args:
        ws: Excel worksheet
        start_row: První řádek sekce (1-indexed, Excel řádkování)
        end_row: Poslední řádek sekce (1-indexed)

    Returns:
        Slovník s číselnými klíči ("1", "2", ...) pro každý řádek sekce
    """
    section_data = {}
    row_counter = 0

    for row_num in range(start_row, end_row + 1):
        row = ws[row_num]

        # Přečti hodnoty ze sloupců A-F (indexy 0-5)
        nazev_polohy = row[COL_NAZEV_POLOHY].value
        typ_prace = row[COL_TYP_PRACE].value
        vyskyt_a = row[COL_VYSKYT_A].value
        vyskyt_b = row[COL_VYSKYT_B].value
        prumer = row[COL_PRUMER].value
        typ_polohy = row[COL_TYP_POLOHY].value

        # Pokud je název polohy prázdný, přeskoč řádek
        if not nazev_polohy:
            continue

        # Zvyš čítač řádků (začíná od 1)
        row_counter += 1

        # Vytvoř slovník pro tento řádek
        row_data = {
            "nazev_polohy": convert_to_json_safe(nazev_polohy),
            "typ_svalove_prace": convert_to_json_safe(typ_prace),
            "vyskyt_min_a": convert_to_json_safe(vyskyt_a),
            "vyskyt_min_b": convert_to_json_safe(vyskyt_b),
            "prumer_min": convert_to_json_safe(prumer),
            "typ_pracovni_polohy": convert_to_json_safe(typ_polohy)
        }

        # Použij číselný klíč jako string
        section_data[str(row_counter)] = row_data

    return section_data


# =============================================================================
# MAIN READER FUNCTION
# =============================================================================

def read_pp_results(
    excel_path: str,
    what_is_evaluated: str = "cas",
    worker_count: int = 2
) -> Dict[str, Any]:
    """
    Hlavní funkce pro načtení PP výsledků z Excelu

    Args:
        excel_path: Cesta k PP Excel souboru (PP_*_CAS.xlsx nebo PP_*_KUSY.xlsx)
        what_is_evaluated: "cas" nebo "kusy" (pro budoucí rozšíření)
        worker_count: Počet pracovníků (1 nebo 2)

    Returns:
        Slovník s výsledky ve struktuře:
        {
            "excel_type": "PP_CAS" nebo "PP_KUSY",
            "worker_count": 2,
            "what_is_evaluated": "cas",
            "trup": [...],
            "hlava_krk": [...],
            "phk": [...],
            "lhk": [...],
            "dk": [...],
            "ostatni": [...]
        }
    """
    excel_path = Path(excel_path)

    # Detekuj typ PP (CAS nebo KUSY)
    if "_CAS" in excel_path.name.upper():
        excel_type = "PP_CAS"
    elif "_KUSY" in excel_path.name.upper():
        excel_type = "PP_KUSY"
    else:
        excel_type = "PP"  # Generic fallback

    print(f"Nacitam PP Excel: {excel_path.name}")
    print(f"Typ: {excel_type}")
    print(f"Pocet pracovniku: {worker_count}")

    # Načti Excel
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Najdi list "Průměr"
    if PRUMER_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"List '{PRUMER_SHEET}' nenalezen v Excel souboru!")

    ws = wb[PRUMER_SHEET]
    print(f"Pouzivam list: Prumer")

    # Inicializuj výsledný slovník
    results = {
        "excel_type": excel_type,
        "worker_count": worker_count,
        "what_is_evaluated": what_is_evaluated
    }

    # Načti data z každé sekce
    for section_name, (start_row, end_row) in SECTIONS.items():
        print(f"  Nacitam sekci: {section_name} (radky {start_row}-{end_row})")

        section_data = read_section_data(ws, start_row, end_row)
        results[section_name] = section_data

        print(f"    OK Nacteno {len(section_data)} polozek")

    wb.close()

    # Přidej metadata
    results["_metadata"] = {
        "source_file": excel_path.name,
        "sheet_name": PRUMER_SHEET,
        "total_sections": len(SECTIONS),
        "total_rows": sum(len(results[section]) for section in SECTIONS.keys())
    }

    print(f"OK PP vysledky nacteny: {results['_metadata']['total_rows']} radku celkem")

    return results


# =============================================================================
# COMMAND LINE INTERFACE
# =============================================================================

def main():
    """CLI interface pro testování"""
    if len(sys.argv) < 2:
        print("Usage: python read_pp_results.py <pp_excel_file>")
        print("\nExample:")
        print("  python read_pp_results.py projects/001-2024_Firma/PP_001-2024_Firma_CAS.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]

    # Detekuj parametry z názvu souboru
    what_is_evaluated = "cas" if "_CAS" in Path(excel_path).name.upper() else "kusy"
    worker_count = 2  # Default

    # Načti výsledky
    try:
        results = read_pp_results(excel_path, what_is_evaluated, worker_count)

        # Ulož do JSON
        output_path = Path(excel_path).parent / "pp_results.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=4)

        print(f"\nOK Vysledky ulozeny do: {output_path}")

        # Zobraz statistiky
        print("\n" + "="*60)
        print("STATISTIKY:")
        print("="*60)
        for section_name in SECTIONS.keys():
            count = len(results[section_name])
            print(f"  {section_name:15} : {count:2} polozek")

    except Exception as e:
        print(f"\nERROR Chyba pri nacitani PP vysledku:")
        print(f"  {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
