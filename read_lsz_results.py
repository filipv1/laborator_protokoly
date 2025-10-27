"""
LSZ Excel Results Reader - načte výsledky z LSZ Excel a uloží do JSON
Usage: python read_lsz_results.py <lsz_excel_file>
"""
import sys
import json
import openpyxl
from datetime import time, datetime
from pathlib import Path
import xlwings as xw


def convert_to_json_safe(value):
    """Převede hodnotu na JSON-safe formát"""
    if isinstance(value, (time, datetime)):
        return str(value)
    return value


# ZAKOMENTOVÁNO: Export grafů z Excelu (xlwings) - funkce odstraněna na žádost uživatele
# def export_charts(excel_path):
#     """Exportuje 2 grafy z LSZ Excel pomocí screenshot přístupu"""
#     excel_path = Path(excel_path)
#     output_dir = excel_path.parent / "lsz_charts"
#     output_dir.mkdir(exist_ok=True)
#
#     app = xw.App(visible=False)
#     wb = app.books.open(str(excel_path))
#     ws = wb.sheets['Celkové výsledky']
#
#     charts = {}
#
#     # Export pomocí Copy -> Paste jako PNG
#     try:
#         from PIL import ImageGrab
#         import win32clipboard
#         from io import BytesIO
#
#         # Graf 1
#         chart1 = ws.charts[0]
#         chart1.api.CopyPicture()
#         img1 = ImageGrab.grabclipboard()
#         if img1:
#             chart1_path = output_dir / "graf1.png"
#             img1.save(str(chart1_path), 'PNG')
#             charts["graf1"] = str(chart1_path)
#             print(f"  ✓ Graf 1 exportován: {chart1_path}")
#
#         # Graf 2
#         chart2 = ws.charts[1]
#         chart2.api.CopyPicture()
#         img2 = ImageGrab.grabclipboard()
#         if img2:
#             chart2_path = output_dir / "graf2.png"
#             img2.save(str(chart2_path), 'PNG')
#             charts["graf2"] = str(chart2_path)
#             print(f"  ✓ Graf 2 exportován: {chart2_path}")
#
#     except Exception as e:
#         print(f"  ✗ Export selhal: {e}")
#
#     wb.close()
#     app.quit()
#
#     return charts

def export_charts(excel_path):
    """ZAKOMENTOVÁNO: Funkce nyní neexportuje grafy - vrací prázdný dict"""
    print("  ℹ Export grafů je zakázán")
    return {}


def read_somatometrie_table(excel_path):
    """Načte tabulku somatometrických dat z listu Somatometrie (oblast B19:H24)"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if 'Somatometrie' not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Somatometrie"]

    # Načti data z oblasti B19:H24
    # Řádek 19 jsou hlavičky, řádky 20-24 jsou data
    somatometrie_data = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=20, max_row=24, min_col=2, max_col=8, values_only=True), start=1):
        # Přeskoč úplně prázdné řádky
        if all(cell is None or cell == '' for cell in row):
            continue

        somatometrie_data[str(row_idx)] = {
            "datum": convert_to_json_safe(row[0]),
            "inicialy": row[1],
            "lateralita": row[2],
            "vek_roky": row[3],
            "expozice_roky": row[4],
            "vyska_cm": row[5],
            "hmotnost_kg": row[6]
        }

    wb.close()
    return somatometrie_data


def read_movements_per_unit_from_worker(wb, sheet_name, what_is_evaluated):
    """
    Načte movements per unit z listu Pracovník A nebo B

    Args:
        wb: Workbook object
        sheet_name: "Pracovník A" nebo "Pracovník B"
        what_is_evaluated: "čas" nebo "kusy"

    Returns:
        List 6 hodnot [row49, row50, row51, row52, row53, row54]
        Každý row je tuple (movements_per_unit_phk, movements_per_unit_lhk)
    """
    if sheet_name not in wb.sheetnames:
        return [(0, 0)] * 6  # Vrátit 6 prázdných řádků

    ws = wb[sheet_name]

    # Podle what_is_evaluated vyber sloupce
    if what_is_evaluated == "čas":
        col_phk = 8   # H
        col_lhk = 9   # I
    else:  # "kusy"
        col_phk = 20  # T
        col_lhk = 21  # U

    # Načti 6 řádků (49-54)
    movements = []
    for row_idx in range(49, 55):  # 49-54 (6 řádků)
        phk_value = ws.cell(row=row_idx, column=col_phk).value or 0
        lhk_value = ws.cell(row=row_idx, column=col_lhk).value or 0
        movements.append((phk_value, lhk_value))

    return movements


def read_lsz_results(excel_path, what_is_evaluated="kusy"):
    """
    Načte výsledky z LSZ Excel souboru

    Args:
        excel_path: Cesta k LSZ Excel souboru
        what_is_evaluated: "čas" nebo "kusy" (z measurement_data.json)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Celkové výsledky"]

    # Single cells
    results = {
        "Fmax_Phk_Extenzor": ws["K45"].value,
        "Fmax_Phk_Flexor": ws["L45"].value,
        "phk_number_of_movements": ws["M45"].value,
        "Fmax_Lhk_Extenzor": ws["O45"].value,
        "Fmax_Lhk_Flexor": ws["Q45"].value,
        "lhk_number_of_movements": ws["S45"].value,
    }

    # Tabulka W4-Y51 (hardcoded)
    table1 = {}
    row_num = 0
    for row in ws.iter_rows(min_row=4, max_row=51, min_col=23, max_col=25, values_only=True):  # W=23, Y=25
        if any(row):  # Přeskoč úplně prázdné řádky
            row_num += 1
            table1[str(row_num)] = {
                "fmax": row[0],
                "phk": row[1],
                "lhk": row[2]
            }
    results["table_W4_Y51"] = table1

    # Tabulka B4-I21 (dynamická - najít podle "Činnost")
    table2_start_row = None

    # Najdi řádek s textem "Činnost" v B sloupci
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, min_col=2, max_col=2, values_only=True), start=1):
        if row[0] and "Činnost" in str(row[0]):
            table2_start_row = row_idx + 3  # Přeskoč 3 řádky headerů
            break

    table2 = {}
    if table2_start_row:
        row_num = 0
        for row in ws.iter_rows(min_row=table2_start_row, max_row=table2_start_row + 50, min_col=2, max_col=9, values_only=True):  # B=2, I=9
            # Break na prázdném řádku (B sloupec prázdný A není číslo 0)
            if not row[0] and row[0] != 0:
                break

            row_num += 1
            table2[str(row_num)] = {
                "activity": row[0],
                "time_min": row[3],
                "phk_extenzory": row[4],
                "phk_flexory": row[5],
                "lhk_extenzory": row[6],
                "lhk_flexory": row[7]
            }

    results["table_B4_I21"] = table2

    # Tabulka 3: Výsledky měřených osob – počet pohybů/jednotka (B28-I41)
    # HARDCODED RANGE: B28-I41 (data začínají od řádku 28)
    table3 = {}
    row_num = 0
    for row in ws.iter_rows(min_row=28, max_row=41, min_col=2, max_col=9, values_only=True):  # B=2, I=9
        # Přeskoč úplně prázdné řádky
        if not row[0] and row[0] != 0:
            continue

        row_num += 1
        table3[str(row_num)] = {
            "activity": row[0],
            "pieces": row[4],
            "time_min": row[5],
            "movements_per_unit_phk": row[6],  # Bude přepsáno z Pracovník A/B
            "movements_per_unit_lhk": row[7],  # Bude přepsáno z Pracovník A/B
            "all_phk": row[6],                 # Z Celkové výsledky H28-H41
            "all_lhk": row[7]                  # Z Celkové výsledky I28-I41
        }

    results["table_movements_per_unit"] = table3

    # POST-PROCESSING: Update indexů 1-6 a 8-13 podle what_is_evaluated
    # Načti movements per unit z listů Pracovník A a Pracovník B
    movements_worker_a = read_movements_per_unit_from_worker(wb, "Pracovník A", what_is_evaluated)
    movements_worker_b = read_movements_per_unit_from_worker(wb, "Pracovník B", what_is_evaluated)

    # Update indexy 1-6 (Pracovník A)
    for i in range(6):
        index = str(i + 1)
        if index in table3:
            phk_val, lhk_val = movements_worker_a[i]
            table3[index]["movements_per_unit_phk"] = phk_val
            table3[index]["movements_per_unit_lhk"] = lhk_val

    # Update indexy 8-13 (Pracovník B)
    for i in range(6):
        index = str(i + 8)  # 8, 9, 10, 11, 12, 13
        if index in table3:
            phk_val, lhk_val = movements_worker_b[i]
            table3[index]["movements_per_unit_phk"] = phk_val
            table3[index]["movements_per_unit_lhk"] = lhk_val

    # Tabulka 4: Výsledky měřených osob – časově vážený průměr (B45-I50)
    table4_start_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, min_col=2, max_col=10, values_only=True), start=1):
        if row[0] and "časově" in str(row[0]).lower() and "průměr" in str(row[0]).lower():
            table4_start_row = row_idx + 4  # Přeskoč header řádky (43 → 47)
            break

    table4 = {}
    if table4_start_row:
        row_num = 0
        for row in ws.iter_rows(min_row=table4_start_row, max_row=table4_start_row + 20, min_col=2, max_col=9, values_only=True):
            # Break pokud je prázdný řádek NEBO začíná další sekce
            if (not row[0] and not row[1]) or (row[0] and "Výsledky" in str(row[0])):
                break

            row_num += 1
            table4[str(row_num)] = {
                "date": convert_to_json_safe(row[0]),
                "worker": convert_to_json_safe(row[1]),
                "phk_extenzory": convert_to_json_safe(row[2]),
                "phk_flexory": convert_to_json_safe(row[3]),
                "phk_movements": convert_to_json_safe(row[4]),
                "lhk_extenzory": convert_to_json_safe(row[5]),
                "lhk_flexory": convert_to_json_safe(row[6]),
                "lhk_movements": convert_to_json_safe(row[7])
            }

    results["table_time_weighted_average"] = table4

    # Tabulka 5: Rozložení vynakládaných svalových sil (B54-J76)
    table5_start_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, min_col=2, max_col=10, values_only=True), start=1):
        if row[0] and "rozložení" in str(row[0]).lower() and "svalových sil" in str(row[0]).lower():
            table5_start_row = row_idx + 4  # Přeskoč header řádky (52 → 56)
            break

    table5 = {}
    if table5_start_row:
        row_num = 0
        for row in ws.iter_rows(min_row=table5_start_row, max_row=table5_start_row + 50, min_col=2, max_col=10, values_only=True):  # B=2, J=10
            # Break na prázdném řádku
            if not row[0] and row[0] != 0:
                break

            row_num += 1
            table5[str(row_num)] = {
                "activity": row[0],
                "force_55_70_phk_extenzory": row[1],
                "force_55_70_phk_flexory": row[2],
                "force_55_70_lhk_extenzory": row[3],
                "force_55_70_lhk_flexory": row[4],
                "force_over_70_phk_extenzory": row[5],
                "force_over_70_phk_flexory": row[6],
                "force_over_70_lhk_extenzory": row[7],
                "force_over_70_lhk_flexory": row[8]
            }

    results["table_force_distribution"] = table5

    wb.close()

    # Přidej somatometrickou tabulku
    somatometrie = read_somatometrie_table(excel_path)
    results["table_somatometrie"] = somatometrie

    return results


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python read_lsz_results.py <lsz_excel_file> <measurement_data_json>")
        sys.exit(1)

    excel_file = sys.argv[1]
    measurement_json = sys.argv[2]

    # Načti measurement_data.json pro zjištění what_is_evaluated
    with open(measurement_json, encoding='utf-8') as f:
        measurement_data = json.load(f)

    what_is_evaluated = measurement_data.get("section3_additional_data", {}).get("what_is_evaluated", "kusy")
    print(f"→ Co se hodnotí: {what_is_evaluated}")

    # Načti výsledky
    results = read_lsz_results(excel_file, what_is_evaluated)

    # ZAKOMENTOVÁNO: Export grafů zakázán
    # # Exportuj grafy
    # print("Exportuji grafy...")
    # charts = export_charts(excel_file)
    # results["charts"] = charts

    # Ulož do JSON
    output_file = "lsz_results.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=4)

    print(f"✓ Výsledky uloženy do: {output_file}")
    # print(f"✓ Grafy exportovány: {len(charts)} obrázků")  # ZAKOMENTOVÁNO
