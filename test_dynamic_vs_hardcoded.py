"""
Test: Ukázka rozdílu mezi HARDCODED a DYNAMICKOU tabulkou
"""
import openpyxl

excel_path = r"templates/excel/LSZ_template.xlsm"
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["Celkové výsledky"]

print("=" * 80)
print("ULTRATHINK: HARDCODED vs DYNAMICKÁ tabulka")
print("=" * 80)
print()

# ============================================
# HARDCODED TABULKA: table_W4_Y51
# ============================================
print("1. HARDCODED TABULKA: table_W4_Y51")
print("-" * 80)
print("Kód:")
print("  for row in ws.iter_rows(min_row=4, max_row=51, min_col=23, max_col=25):")
print()

hardcoded_count = 0
for row in ws.iter_rows(min_row=4, max_row=51, min_col=23, max_col=25, values_only=True):
    if any(row):
        hardcoded_count += 1

print(f"Výsledek: Načteno {hardcoded_count} řádků (vždy same rozsah 4-51)")
print()
print("❌ PROBLÉM: Pokud Excel má:")
print("   - Více řádků (např. 60) → odřízne data po 51. řádku")
print("   - Méně řádků (např. 30) → načte i prázdné řádky 31-51")
print()

# ============================================
# DYNAMICKÁ TABULKA: table_B4_I21
# ============================================
print("2. DYNAMICKÁ TABULKA: table_B4_I21")
print("-" * 80)
print("Kód:")
print('  # KROK 1: Najdi začátek (hledá "Činnost")')
print('  for row in ws.iter_rows(min_row=1, max_row=30, min_col=2, max_col=2):')
print('      if "Činnost" in str(row[0]):')
print('          start_row = row_idx + 3')
print()
print("  # KROK 2: Čti až do prázdného řádku")
print("  for row in ws.iter_rows(min_row=start_row, max_row=start_row+50, ...):")
print("      if not row[0] and row[0] != 0:")
print("          break  # <--- ZASTAVÍ SE NA PRÁZDNÉM ŘÁDKU!")
print()

# Simuluj dynamickou logiku
table_start_row = None
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, min_col=2, max_col=2, values_only=True), start=1):
    if row[0] and "Činnost" in str(row[0]):
        table_start_row = row_idx + 3
        print(f"✓ Našel nadpis 'Činnost' na řádku {row_idx}")
        print(f"✓ Začátek dat: řádek {table_start_row}")
        break

if table_start_row:
    dynamic_count = 0
    for row in ws.iter_rows(min_row=table_start_row, max_row=table_start_row + 50, min_col=2, max_col=9, values_only=True):
        if not row[0] and row[0] != 0:
            print(f"✓ Zastavil se na prázdném řádku (řádek {table_start_row + dynamic_count})")
            break
        dynamic_count += 1

    print(f"✓ Načteno {dynamic_count} řádků dat")
    print()
    print("✅ VÝHODA: Automaticky se přizpůsobí:")
    print("   - Pokud Excel má 5 řádků → načte 5")
    print("   - Pokud Excel má 15 řádků → načte 15")
    print("   - Pokud Excel má 80 řádků → načte 50 (max limit)")
    print()

# ============================================
# SROVNÁNÍ
# ============================================
print("=" * 80)
print("SROVNÁNÍ")
print("=" * 80)
print(f"HARDCODED (table_W4_Y51):     {hardcoded_count} řádků (fixní rozsah)")
print(f"DYNAMICKÁ (table_B4_I21):     {dynamic_count} řádků (automatická detekce)")
print()
print("ZÁVĚR:")
print("  - 'table_B4_I21' JE dynamická (špatný název, dobrá logika)")
print("  - 'table_W4_Y51' JE hardcoded (dobrý název, špatná logika)")
print("  - 'table_somatometrie' JE hardcoded (B19:H24 fixní)")
print()

wb.close()
