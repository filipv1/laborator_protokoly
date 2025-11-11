"""
Test opravy kopírování kusů do LSZ Excelu.

Testuje 3 scénáře:
1. Čas + Kusy (oba jsou vyplněné)
2. Jen čas (kusy chybí)
3. Jen kusy (čas chybí)
"""
from pathlib import Path
from core.table_copier import TableCopier
import openpyxl

def print_excel_values(excel_path: Path, row: int):
    """Vypise hodnoty z daneho radku LSZ Excelu."""
    wb = openpyxl.load_workbook(excel_path, keep_vba=True)
    ws = wb["Časový snímek"]

    operation = ws[f"C{row}"].value
    time_min = ws[f"F{row}"].value
    pieces = ws[f"J{row}"].value

    # Encode-safe print
    operation_safe = operation.encode('ascii', 'ignore').decode('ascii') if operation else None
    print(f"  Radek {row}: C='{operation_safe}', F={time_min}, J={pieces}")
    wb.close()


def test_scenario_1_both():
    """Scenar 1: Cas + Kusy"""
    print("\n" + "="*70)
    print("SCENAR 1: CAS + KUSY (oba vyplnene)")
    print("="*70)

    time_schedule = {
        "line1": {
            "operation": "Montáž součástky",
            "time_min": 45,
            "pieces_count": 120
        }
    }

    # Najdi LSZ Excel sablonu
    template_path = Path("templates/excel/LSZ_template.xlsm")
    if not template_path.exists():
        print(f"CHYBA: Sablona {template_path} neexistuje!")
        return

    # Vytvoř kopii pro test
    test_path = Path("test_lsz_scenario1.xlsm")
    import shutil
    shutil.copy(template_path, test_path)

    # Zkopíruj data
    copier = TableCopier()
    copier.copy_time_schedule_to_lsz(test_path, time_schedule)

    # Ověř výsledek
    print("\nOcekavany vysledek:")
    print("  C26: 'Montaz soucastky'")
    print("  F26: 45")
    print("  J26: 120 (ne vypocitana norma!)")

    print("\nSkutecny vysledek:")
    print_excel_values(test_path, 26)

    # Cleanup
    test_path.unlink()


def test_scenario_2_time_only():
    """Scenar 2: Jen cas"""
    print("\n" + "="*70)
    print("SCENAR 2: JEN CAS (kusy chybi)")
    print("="*70)

    time_schedule = {
        "line1": {
            "operation": "Příprava materiálu",
            "time_min": 30,
            "pieces_count": None  # Chybí!
        }
    }

    template_path = Path("templates/excel/LSZ_template.xlsm")
    if not template_path.exists():
        print(f"CHYBA: Sablona {template_path} neexistuje!")
        return

    test_path = Path("test_lsz_scenario2.xlsm")
    import shutil
    shutil.copy(template_path, test_path)

    copier = TableCopier()
    copier.copy_time_schedule_to_lsz(test_path, time_schedule)

    print("\nOcekavany vysledek:")
    print("  C26: 'Priprava materialu'")
    print("  F26: 30")
    print("  J26: None nebo prazdne (kusy chybi)")

    print("\nSkutecny vysledek:")
    print_excel_values(test_path, 26)

    test_path.unlink()


def test_scenario_3_pieces_only():
    """Scenar 3: Jen kusy"""
    print("\n" + "="*70)
    print("SCENAR 3: JEN KUSY (cas chybi)")
    print("="*70)

    time_schedule = {
        "line1": {
            "operation": "Balení produktů",
            "time_min": None,  # Chybí!
            "pieces_count": 200
        }
    }

    template_path = Path("templates/excel/LSZ_template.xlsm")
    if not template_path.exists():
        print(f"CHYBA: Sablona {template_path} neexistuje!")
        return

    test_path = Path("test_lsz_scenario3.xlsm")
    import shutil
    shutil.copy(template_path, test_path)

    copier = TableCopier()
    copier.copy_time_schedule_to_lsz(test_path, time_schedule)

    print("\nOcekavany vysledek:")
    print("  C26: 'Baleni produktu'")
    print("  F26: None nebo prazdne (cas chybi)")
    print("  J26: 200 (KUSY SE VYPISU!)")

    print("\nSkutecny vysledek:")
    print_excel_values(test_path, 26)

    test_path.unlink()


if __name__ == "__main__":
    print("\nTEST OPRAVY KOPIROVANI KUSU DO LSZ EXCELU")
    print("=" * 70)

    test_scenario_1_both()
    test_scenario_2_time_only()
    test_scenario_3_pieces_only()

    print("\n" + "="*70)
    print("VSECHNY TESTY DOKONCENY")
    print("="*70)
