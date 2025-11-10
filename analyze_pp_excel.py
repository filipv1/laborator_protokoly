"""
Quick script to analyze PP Excel structure - output to file
"""
import openpyxl
from pathlib import Path

# Load PP Excel
excel_path = Path("projects/_fjfjfe/PP__fjfjfe_CAS.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Write to file
output_file = Path("PP_EXCEL_ANALYSIS.txt")
lines = []

lines.append("="*60)
lines.append("PP EXCEL STRUCTURE ANALYSIS")
lines.append("="*60)
lines.append("")

lines.append(f"File: {excel_path.name}")
lines.append(f"Sheets: {wb.sheetnames}")

# Find 'Prumer' sheet
prumer_sheet = None
for sheet_name in wb.sheetnames:
    if 'Pr' in sheet_name and 'm' in sheet_name:  # Prumer
        prumer_sheet = sheet_name
        break

if not prumer_sheet:
    lines.append("ERROR: 'Prumer' sheet not found!")
    prumer_sheet = wb.sheetnames[-2]  # Try second to last

ws = wb[prumer_sheet]
lines.append(f"Using sheet: '{ws.title}'")
lines.append("")

lines.append("="*60)
lines.append("TABULKA STRUKTURY:")
lines.append("="*60)

# Define all sections
sections = [
    ("Trup", 4, 14),
    ("Hlava a krk", 16, 25),
    ("PHK", 27, 36),
    ("LHK", 38, 47),
    ("DK", 49, 55),
    ("Ostatni", 57, 60)
]

# Print header row (assume row 3 is header)
lines.append("")
lines.append("Header row (radek 3):")
header_row = ws[3]
for col_idx, cell in enumerate(header_row[0:6], start=1):  # A-F
    col_letter = chr(64 + col_idx)  # A=65
    lines.append(f"  {col_letter}: {cell.value}")

# Print sample from each section
for section_name, start_row, end_row in sections:
    lines.append("")
    lines.append(f"{section_name} (radky {start_row}-{end_row}):")

    # Print first 3 rows of section
    for sample_idx in range(3):
        row_num = start_row + sample_idx
        if row_num > end_row:
            break

        row = ws[row_num]
        lines.append(f"  Radek {row_num}:")
        for col_idx, cell in enumerate(row[0:6], start=1):
            col_letter = chr(64 + col_idx)
            value = cell.value
            # Format value
            if isinstance(value, (int, float)):
                value = f"{value}"
            elif value is None:
                value = "[EMPTY]"
            lines.append(f"    {col_letter}: {value}")

    # Count non-empty rows
    non_empty = 0
    for row_idx in range(start_row, end_row + 1):
        row = ws[row_idx]
        if row[0].value:  # Column A has value
            non_empty += 1

    lines.append(f"  >> Pocet neprazdnych radku: {non_empty}/{end_row - start_row + 1}")

wb.close()

lines.append("")
lines.append("="*60)
lines.append("ANALYSIS COMPLETE")
lines.append("="*60)

# Write to file
with open(output_file, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Analysis saved to: {output_file}")
print(f"Total lines: {len(lines)}")
