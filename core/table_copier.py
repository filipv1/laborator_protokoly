"""
Table Copier - kopírování tabulkových dat do Excelů
"""
import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional
from config.table_mappings import (
    CFZ_TABLE_MAPPING,
    LSZ_TABLE_MAPPING,
    PP_CAS_TABLE_MAPPING,
    PP_KUSY_TABLE_MAPPING
)


class TableCopier:
    """Kopírování časového snímku do Excel souborů"""

    def copy_time_schedule_to_cfz(self, excel_path: Path, time_schedule: Dict[str, Any]) -> None:
        """
        Zkopíruje časový snímek do CFZ Excel souboru.

        Args:
            excel_path: Cesta k CFZ Excel souboru
            time_schedule: Data časového snímku (line1...lineN + total)
        """
        mapping = CFZ_TABLE_MAPPING
        wb = openpyxl.load_workbook(excel_path, keep_vba=False)

        if mapping["sheet"] not in wb.sheetnames:
            print(f"Varování: List '{mapping['sheet']}' neexistuje v {excel_path.name}")
            wb.close()
            return

        ws = wb[mapping["sheet"]]

        # Zkopíruj data řádek po řádku
        row_num = mapping["start_row"]
        for i in range(1, 21):  # Max 20 řádků
            line_key = f"line{i}"
            if line_key not in time_schedule:
                break

            line_data = time_schedule[line_key]

            # Přeskoč prázdné řádky
            if not line_data.get("operation") and not line_data.get("time_min"):
                continue

            # Vyplň buňky
            ws[f"{mapping['columns']['operation']}{row_num}"] = line_data.get("operation", "")
            ws[f"{mapping['columns']['time_min']}{row_num}"] = line_data.get("time_min")
            # Poznámky necháme prázdné (uživatel může vyplnit manuálně)

            row_num += 1

        wb.save(excel_path)
        wb.close()
        print(f"  ✓ Časový snímek zkopírován do CFZ ({row_num - mapping['start_row']} řádků)")

    def copy_time_schedule_to_lsz(self, excel_path: Path, time_schedule: Dict[str, Any]) -> None:
        """
        Zkopíruje časový snímek do LSZ Excel souboru.

        Args:
            excel_path: Cesta k LSZ Excel souboru
            time_schedule: Data časového snímku
        """
        mapping = LSZ_TABLE_MAPPING
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)  # LSZ má makra!

        if mapping["sheet"] not in wb.sheetnames:
            print(f"Varování: List '{mapping['sheet']}' neexistuje v {excel_path.name}")
            wb.close()
            return

        ws = wb[mapping["sheet"]]

        # Zkopíruj data
        row_num = mapping["start_row"]
        for i in range(1, 21):
            line_key = f"line{i}"
            if line_key not in time_schedule:
                break

            line_data = time_schedule[line_key]

            if not line_data.get("operation") and not line_data.get("time_min"):
                continue

            # Vyplň buňky
            ws[f"{mapping['columns']['operation']}{row_num}"] = line_data.get("operation", "")
            ws[f"{mapping['columns']['time_min']}{row_num}"] = line_data.get("time_min")

            # Norma [ks/hod] - vypočítej z pieces_count a time_min
            pieces = line_data.get("pieces_count")
            time_min = line_data.get("time_min")
            if pieces and time_min and time_min > 0:
                # ks/hod = (počet kusů / čas v min) * 60
                norm = (pieces / time_min) * 60
                ws[f"{mapping['columns']['norm_pcs_hour']}{row_num}"] = round(norm, 2)

            row_num += 1

        wb.save(excel_path)
        wb.close()
        print(f"  ✓ Časový snímek zkopírován do LSZ ({row_num - mapping['start_row']} řádků)")

    def copy_time_schedule_to_pp_cas(self, excel_path: Path, time_schedule: Dict[str, Any]) -> None:
        """
        Zkopíruje časový snímek do PP ČAS Excel souboru.

        Args:
            excel_path: Cesta k PP ČAS Excel souboru
            time_schedule: Data časového snímku
        """
        mapping = PP_CAS_TABLE_MAPPING
        wb = openpyxl.load_workbook(excel_path, keep_vba=False)

        if mapping["sheet"] not in wb.sheetnames:
            print(f"Varování: List '{mapping['sheet']}' neexistuje v {excel_path.name}")
            wb.close()
            return

        ws = wb[mapping["sheet"]]

        # Zkopíruj data (max 30 řádků)
        row_num = mapping["start_row"]
        for i in range(1, 31):  # PP má 30 řádků
            line_key = f"line{i}"
            if line_key not in time_schedule:
                break

            line_data = time_schedule[line_key]

            if not line_data.get("operation") and not line_data.get("time_min"):
                continue

            # Vyplň buňky
            ws[f"{mapping['columns']['operation']}{row_num}"] = line_data.get("operation", "")

            # Konverze minut na sekundy
            time_min = line_data.get("time_min")
            if time_min:
                time_sec = time_min * 60
                ws[f"{mapping['columns']['time_sec']}{row_num}"] = time_sec

            row_num += 1

        wb.save(excel_path)
        wb.close()
        print(f"  ✓ Časový snímek zkopírován do PP ČAS ({row_num - mapping['start_row']} řádků)")

    def copy_time_schedule_to_pp_kusy(self, excel_path: Path, time_schedule: Dict[str, Any]) -> None:
        """
        Zkopíruje časový snímek do PP KUSY Excel souboru.

        Args:
            excel_path: Cesta k PP KUSY Excel souboru
            time_schedule: Data časového snímku
        """
        mapping = PP_KUSY_TABLE_MAPPING
        wb = openpyxl.load_workbook(excel_path, keep_vba=False)

        if mapping["sheet"] not in wb.sheetnames:
            print(f"Varování: List '{mapping['sheet']}' neexistuje v {excel_path.name}")
            wb.close()
            return

        ws = wb[mapping["sheet"]]

        # Zkopíruj data (max 30 řádků)
        row_num = mapping["start_row"]
        for i in range(1, 31):
            line_key = f"line{i}"
            if line_key not in time_schedule:
                break

            line_data = time_schedule[line_key]

            if not line_data.get("operation") and not line_data.get("time_min"):
                continue

            # Vyplň buňky
            ws[f"{mapping['columns']['operation']}{row_num}"] = line_data.get("operation", "")

            # Konverze minut na sekundy
            time_min = line_data.get("time_min")
            if time_min:
                time_sec = time_min * 60
                ws[f"{mapping['columns']['time_sec']}{row_num}"] = time_sec

            row_num += 1

        wb.save(excel_path)
        wb.close()
        print(f"  ✓ Časový snímek zkopírován do PP KUSY ({row_num - mapping['start_row']} řádků)")

    def copy_time_schedule(self, excel_path: Path, excel_type: str, time_schedule: Dict[str, Any]) -> None:
        """
        Univerzální metoda pro kopírování časového snímku podle typu Excelu.

        Args:
            excel_path: Cesta k Excel souboru
            excel_type: Typ Excel souboru ("lsz", "pp_time", "pp_pieces", "cfz")
            time_schedule: Data časového snímku
        """
        if excel_type == "lsz":
            self.copy_time_schedule_to_lsz(excel_path, time_schedule)
        elif excel_type == "pp_time":
            self.copy_time_schedule_to_pp_cas(excel_path, time_schedule)
        elif excel_type == "pp_pieces":
            self.copy_time_schedule_to_pp_kusy(excel_path, time_schedule)
        elif excel_type == "cfz":
            self.copy_time_schedule_to_cfz(excel_path, time_schedule)
        else:
            print(f"Varování: Neznámý typ Excel souboru: {excel_type}")
