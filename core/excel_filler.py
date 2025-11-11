"""
Excel Filler - vyplňování dat do Excel souborů
"""
import openpyxl
from pathlib import Path
from typing import Dict, Any


class ExcelFiller:
    """Vyplňování dat z JSON do Excel šablon"""

    def __init__(self, mapping: Dict[str, Dict[str, str]]):
        """
        Args:
            mapping: Dictionary s mappingem (např. LSZ_MAPPING)
                     Struktura: {"Název_listu": {"A1": "json.path", ...}}
        """
        self.mapping = mapping

    def fill_excel(self, excel_path: Path, data: Dict[str, Any]) -> None:
        """
        Vyplní Excel soubor daty z JSON.

        Args:
            excel_path: Cesta k Excel souboru
            data: Data z JSON (structure z measurement_data.json)
        """
        # Načti Excel - keep_vba=True jen pro .xlsm (soubory s makry)
        is_macro_file = excel_path.suffix.lower() == '.xlsm'
        wb = openpyxl.load_workbook(excel_path, keep_vba=is_macro_file)

        # Projdi všechny listy v mappingu
        for sheet_name, cell_mapping in self.mapping.items():
            if sheet_name not in wb.sheetnames:
                print(f"Varování: List '{sheet_name}' neexistuje v {excel_path.name}")
                continue

            ws = wb[sheet_name]

            # Vyplň jednotlivé buňky
            for cell_address, json_path in cell_mapping.items():
                value = self._get_value_from_json(data, json_path)

                if value is not None:
                    ws[cell_address] = value

        # Uloží zpět (zachová makra)
        wb.save(excel_path)

    def _get_value_from_json(self, data: Dict[str, Any], path: str) -> Any:
        """
        Získá hodnotu z JSON podle tečkové cesty nebo provede výpočet.

        Args:
            data: JSON data
            path: Tečková cesta (např. "section3_worker_a.full_name")
                  nebo výpočet (např. "CALC:section4_worker_a.work_duration-section4_worker_a.breaks")

        Returns:
            Hodnota nebo None pokud nenalezena
        """
        # Pokud začíná "CALC:", jedná se o výpočet
        if path.startswith("CALC:"):
            return self._calculate_value(data, path[5:])  # Odstraň "CALC:" prefix

        # Jinak normální cesta
        keys = path.split('.')
        current = data

        try:
            for key in keys:
                current = current[key]
            return current
        except (KeyError, TypeError):
            return None

    def _calculate_value(self, data: Dict[str, Any], expression: str) -> Any:
        """
        Provede matematický výpočet z JSON hodnot.

        Args:
            data: JSON data
            expression: Výraz (např. "section4_worker_a.work_duration-section4_worker_a.breaks")

        Returns:
            Vypočtená hodnota nebo None pokud se nepodaří
        """
        try:
            # Najdi operátor (zatím podporujeme jen odčítání)
            if '-' in expression:
                parts = expression.split('-')
                if len(parts) == 2:
                    left_path = parts[0].strip()
                    right_path = parts[1].strip()

                    # Získej hodnoty
                    left_value = self._get_value_from_json(data, left_path)
                    right_value = self._get_value_from_json(data, right_path)

                    # Převeď na čísla a vypočti
                    if left_value is not None and right_value is not None:
                        left_num = float(left_value) if not isinstance(left_value, (int, float)) else left_value
                        right_num = float(right_value) if not isinstance(right_value, (int, float)) else right_value
                        return left_num - right_num

            return None
        except (ValueError, TypeError, KeyError):
            return None
